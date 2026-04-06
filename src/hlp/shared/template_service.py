"""Pricing template service: upload, data-validation extraction, mapping updates."""

from __future__ import annotations

import logging
from pathlib import Path, PurePosixPath
from typing import Any
from uuid import uuid4

from sqlalchemy.orm import Session

from hlp.config import get_settings
from hlp.models.pricing_template import PricingTemplate
from hlp.repositories import pricing_template_repository
from hlp.shared.exceptions import InvalidTemplateError, TemplateNotFoundError

logger = logging.getLogger(__name__)

CATEGORY_PRICING_TEMPLATES = "pricing-templates"

_ALLOWED_EXTENSIONS = {".xlsx", ".xlsm"}


def _validate_extension(filename: str) -> str:
    """Return lowercased extension if allowed, otherwise raise."""
    ext = Path(filename).suffix.lower()
    if ext not in _ALLOWED_EXTENSIONS:
        raise InvalidTemplateError(
            f"Unsupported file type '{ext}'. Allowed: {', '.join(sorted(_ALLOWED_EXTENSIONS))}"
        )
    return ext


def _save_template_file(brand: str, filename: str, content: bytes) -> str:
    """Save file to storage volume and return relative POSIX path."""
    ext = _validate_extension(filename)
    safe_brand = brand.lower().replace(" ", "_")
    stored_name = f"{safe_brand}_{uuid4().hex}{ext}"
    base_path = Path(get_settings().storage_base_path)
    directory = base_path / CATEGORY_PRICING_TEMPLATES
    directory.mkdir(parents=True, exist_ok=True)
    abs_path = directory / stored_name
    abs_path.write_bytes(content)
    return str(PurePosixPath(CATEGORY_PRICING_TEMPLATES) / stored_name)


def upload_template(
    db: Session, brand: str, file_content: bytes, filename: str
) -> PricingTemplate:
    """Upload (or replace) a pricing template for a brand.

    Saves the file to Railway Volume and stores the path in DB.
    If a template already exists for this brand, the file is replaced
    but existing mappings are preserved.
    """
    _validate_extension(filename)
    stored_path = _save_template_file(brand, filename, file_content)

    existing = pricing_template_repository.get_by_brand(db, brand)
    if existing is not None:
        # Replace file but keep mappings
        existing.file_path = stored_path
        # Re-extract data validations from the new file
        validations = extract_data_validations(stored_path)
        existing.data_validations = validations
        db.flush()
        return existing

    # New template — set defaults
    validations = extract_data_validations(stored_path)
    template = pricing_template_repository.create(
        db,
        brand=brand,
        file_path=stored_path,
        sheet_name="Sheet1",
        data_start_row=2,
        header_mappings={},
        column_mappings={},
        data_validations=validations,
    )
    return template


def extract_data_validations(file_path: str) -> dict[str, list[str]]:
    """Extract data validation dropdown options from an Excel template.

    Uses openpyxl to read data validations from the uploaded file.
    Returns a dict like ``{"house_type": ["Access 18", ...], "facade_type": ["Barton", ...]}``.
    """
    from openpyxl import load_workbook

    base_path = Path(get_settings().storage_base_path)
    rel = PurePosixPath(file_path)
    abs_path = base_path.joinpath(*rel.parts)

    if not abs_path.exists():
        logger.warning("Template file not found at %s", abs_path)
        return {}

    try:
        wb = load_workbook(str(abs_path), data_only=True)
    except Exception as exc:
        logger.warning("Could not open template file %s: %s", abs_path, exc)
        return {}

    validations: dict[str, list[str]] = {}
    ws = wb.active
    if ws is None:
        return validations

    # Known field-name patterns to search for in data validations
    field_patterns = {
        "house_type": "house_type",
        "facade_type": "facade_type",
        "bdm": "bdm",
        "wholesale_group": "wholesale_group",
        "garage_side": "garage_side",
    }

    if hasattr(ws, "data_validations") and ws.data_validations:
        for dv in ws.data_validations.dataValidation:
            if dv.type == "list" and dv.formula1:
                formula = str(dv.formula1).strip('"')
                values = [v.strip() for v in formula.split(",") if v.strip()]
                if not values:
                    continue
                # Try to determine the field name from the cell(s) the validation applies to
                field_name = _guess_field_name(ws, dv, field_patterns)
                if field_name:
                    validations[field_name] = values

    wb.close()
    return validations


def _guess_field_name(
    ws: Any, dv: Any, patterns: dict[str, str]
) -> str | None:
    """Try to determine which field a data validation applies to.

    Checks the header cell (row 1) above the validation range for known field names.
    Falls back to the cell reference itself.
    """
    try:
        # Get first cell in validation range
        sqref = str(dv.sqref)
        cell_ref = sqref.split(":")[0].split(" ")[0]
        from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

        col_letter, _row = coordinate_from_string(cell_ref)
        col_idx = column_index_from_string(col_letter)

        # Check header (row 1) for a pattern match
        header_val = ws.cell(row=1, column=col_idx).value
        if header_val:
            header_lower = str(header_val).lower().replace(" ", "_")
            for pattern, field_name in patterns.items():
                if pattern in header_lower:
                    return field_name
            # If no pattern match, use the header value as-is
            return header_lower

    except Exception:
        pass

    return None


def update_template_mappings(
    db: Session,
    template_id: int,
    sheet_name: str | None = None,
    data_start_row: int | None = None,
    header_mappings: dict[str, Any] | None = None,
    column_mappings: dict[str, Any] | None = None,
) -> PricingTemplate:
    """Update cell mappings on an existing template."""
    template = pricing_template_repository.get_by_id(db, template_id)
    if template is None:
        raise TemplateNotFoundError(f"Pricing template {template_id} not found")

    fields: dict[str, Any] = {}
    if sheet_name is not None:
        fields["sheet_name"] = sheet_name
    if data_start_row is not None:
        fields["data_start_row"] = data_start_row
    if header_mappings is not None:
        fields["header_mappings"] = header_mappings
    if column_mappings is not None:
        fields["column_mappings"] = column_mappings

    if fields:
        pricing_template_repository.update(db, template, **fields)

    return template


def get_template_for_brand(db: Session, brand: str) -> PricingTemplate | None:
    """Get the pricing template for a brand (returns None if not uploaded yet)."""
    return pricing_template_repository.get_by_brand(db, brand)
