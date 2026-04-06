"""Pricing request service: submit, fulfil, estimate, list, resubmit, delete."""

from __future__ import annotations

import io
import logging
import math
from datetime import datetime, timezone
from decimal import Decimal
from typing import Any

from sqlalchemy import select
from sqlalchemy.orm import Session

from hlp.api.schemas.common import PaginatedResponse
from hlp.api.schemas.estimator import EstimatorSubmission
from hlp.api.schemas.pricing_request import (
    PricingRequestCreate,
    PricingRequestDetailRead,
    PricingRequestRead,
)
from hlp.models.enums import PricingRequestStatus, UserRoleType
from hlp.models.estate import Estate
from hlp.models.estate_stage import EstateStage
from hlp.models.house_package import HousePackage
from hlp.models.pricing_request import PricingRequest
from hlp.models.profile import Profile
from hlp.models.stage_lot import StageLot
from hlp.repositories import pricing_request_repository
from hlp.shared import notification_service
from hlp.shared.clash_service import get_restricted_lots_for, validate_clash_submission
from hlp.shared.exceptions import (
    ClashViolationError,
    NotAuthorizedError,
)
from hlp.shared.pricing_engine import (
    LotPricingInput,
    PricingContext,
    calculate_batch,
)
from hlp.shared.spreadsheet_service import generate_pricing_spreadsheet
from hlp.shared.storage_service import get_storage_service
from hlp.shared.template_service import get_template_for_brand

logger = logging.getLogger(__name__)


def _has_role(profile: Profile, role: UserRoleType) -> bool:
    return any(ur.role == role for ur in profile.user_roles)


def _is_admin_or_pricing(profile: Profile) -> bool:
    return _has_role(profile, UserRoleType.ADMIN) or _has_role(profile, UserRoleType.PRICING)


def submit_pricing_request(
    db: Session,
    requester_id: int,
    payload: PricingRequestCreate,
) -> PricingRequest:
    """Submit a new pricing request with clash validation and spreadsheet generation."""

    # 1. Brand-specific validation is handled by schema model_validator

    # 2. Two-pass clash validation
    submissions = [
        {
            "lot_number": lot.lot_number,
            "design": lot.house_type,
            "facade": lot.facade_type,
        }
        for lot in payload.lots
    ]

    # Pass 1 — within request
    within_violations = validate_clash_submission(
        db, payload.estate_id, payload.stage_id, submissions
    )

    # Pass 2 — against existing packages
    existing_violations = _check_existing_packages(
        db, payload.estate_id, payload.stage_id, submissions
    )

    all_violations = []
    for v in within_violations:
        all_violations.append({**v, "violation_type": "within_request"})
    for v in existing_violations:
        all_violations.append({**v, "violation_type": "existing_package"})

    if all_violations:
        raise ClashViolationError(
            "Clash rule violations detected", violations=all_violations
        )

    # 3. Create PricingRequest record
    form_data = payload.model_dump(mode="json")
    lot_numbers = [lot.lot_number for lot in payload.lots]

    request = pricing_request_repository.create(
        db,
        requester_id=requester_id,
        estate_id=payload.estate_id,
        stage_id=payload.stage_id,
        brand=payload.brand,
        status=PricingRequestStatus.PENDING,
        form_data=form_data,
        lot_numbers=lot_numbers,
        submitted_at=datetime.now(timezone.utc),
    )
    db.flush()

    # Spreadsheet generation now happens after estimator submits site costs
    # (in submit_estimate). The request stays in Pending status until an
    # estimator is assigned.

    return request


def _check_existing_packages(
    db: Session,
    estate_id: int,
    stage_id: int,
    submissions: list[dict],
) -> list[dict]:
    """Pass 2 clash validation: check submitted lots against existing packages."""
    violations: list[dict] = []

    for sub in submissions:
        lot_number = str(sub["lot_number"])
        design = str(sub.get("design", "")).strip().lower()
        facade = str(sub.get("facade", "")).strip().lower()
        if not design or not facade:
            continue

        restricted_lots = get_restricted_lots_for(db, estate_id, stage_id, lot_number)
        if not restricted_lots:
            continue

        # Check existing packages on restricted lots
        for restricted_lot in restricted_lots:
            stmt = select(HousePackage).where(
                HousePackage.estate_id == estate_id,
                HousePackage.stage_id == stage_id,
                HousePackage.lot_number == restricted_lot,
            )
            existing_packages = list(db.execute(stmt).scalars().all())
            for pkg in existing_packages:
                if (
                    pkg.design.strip().lower() == design
                    and pkg.facade.strip().lower() == facade
                ):
                    violations.append(
                        {
                            "lot_numbers": sorted([lot_number, restricted_lot]),
                            "design": sub["design"],
                            "facade": sub["facade"],
                            "rule_id": None,
                        }
                    )

    return violations


def fulfil_pricing_request(
    db: Session,
    request_id: int,
    file_content: bytes,
    filename: str,
) -> PricingRequest:
    """Upload completed sheet, extract packages, mark as completed."""
    request = pricing_request_repository.get_or_raise(db, request_id)

    if request.status == PricingRequestStatus.COMPLETED:
        raise ValueError("Request is already completed")
    if request.status not in (
        PricingRequestStatus.PRICED,
        PricingRequestStatus.IN_PROGRESS,
        PricingRequestStatus.PENDING,  # backward compat
    ):
        raise ValueError(f"Cannot fulfil request in '{request.status.value}' status")

    # 1. Save uploaded file
    storage = get_storage_service()
    stored_path, _ = storage.save_file("generated-sheets", filename, file_content)

    # 2. Extract packages from Excel and create HousePackage records
    _extract_and_import_packages(db, request, file_content)

    # 3. Update request status
    pricing_request_repository.update(
        db,
        request,
        status=PricingRequestStatus.COMPLETED,
        completed_at=datetime.now(timezone.utc),
        completed_file_path=stored_path,
    )

    # 4. Create notification for requester
    notification_service.create_notification(
        db,
        profile_id=request.requester_id,
        pricing_request_id=request.request_id,
        title="Pricing Request Completed",
        message=f"Your pricing request #{request.request_id} has been completed. "
        f"Download the completed spreadsheet to review packages.",
    )

    return request


def _extract_and_import_packages(
    db: Session,
    request: PricingRequest,
    file_content: bytes,
) -> None:
    """Read lot rows from completed sheet and create HousePackage records."""
    from openpyxl import load_workbook

    template = get_template_for_brand(db, request.brand)
    if template is None:
        logger.warning("No template found for brand %s during extraction", request.brand)
        return

    try:
        wb = load_workbook(io.BytesIO(file_content), data_only=True)
    except Exception as exc:
        logger.warning("Could not open completed sheet: %s", exc)
        return

    if template.sheet_name in wb.sheetnames:
        ws = wb[template.sheet_name]
    else:
        ws = wb.active

    column_mappings = template.column_mappings or {}
    data_start_row = template.data_start_row

    # Read rows until we find an empty lot_number
    lot_number_col = column_mappings.get("lot_number")
    design_col = column_mappings.get("house_type")
    facade_col = column_mappings.get("facade_type")

    if not all([lot_number_col, design_col, facade_col]):
        logger.warning("Missing column mappings for package extraction")
        wb.close()
        return

    row = data_start_row
    while row <= ws.max_row:
        lot_number = ws.cell(row=row, column=lot_number_col).value
        design = ws.cell(row=row, column=design_col).value
        facade = ws.cell(row=row, column=facade_col).value

        if not lot_number:
            break

        lot_number = str(lot_number).strip()
        if not design or not facade:
            row += 1
            continue

        design = str(design).strip()
        facade = str(facade).strip()

        pkg = HousePackage(
            estate_id=request.estate_id,
            stage_id=request.stage_id,
            lot_number=lot_number,
            design=design,
            facade=facade,
            brand=request.brand,
            source=f"pricing_request_{request.request_id}",
        )
        db.add(pkg)
        row += 1

    db.flush()
    wb.close()


def assign_estimator(
    db: Session,
    request_id: int,
    estimator_id: int,
    assigner_profile: Profile,
) -> PricingRequest:
    """Assign an estimator to a pending pricing request."""
    request = pricing_request_repository.get_or_raise(db, request_id)

    if request.status != PricingRequestStatus.PENDING:
        raise ValueError(
            f"Can only assign estimator to Pending requests, "
            f"current status is '{request.status.value}'"
        )

    # Validate the assigner is admin or pricing
    if not _is_admin_or_pricing(assigner_profile):
        raise NotAuthorizedError("Only admin or pricing users can assign estimators")

    # Validate estimator has pricing or admin role
    estimator = db.get(Profile, estimator_id)
    if estimator is None:
        raise ValueError(f"Estimator profile {estimator_id} not found")
    estimator_roles = {ur.role for ur in estimator.user_roles}
    if not (estimator_roles & {UserRoleType.PRICING, UserRoleType.ADMIN}):
        raise ValueError("Estimator must have pricing or admin role")

    pricing_request_repository.update(
        db,
        request,
        estimator_id=estimator_id,
        status=PricingRequestStatus.ESTIMATING,
    )
    return request


def submit_estimate(
    db: Session,
    request_id: int,
    submission: EstimatorSubmission,
    estimator_profile: Profile,
) -> PricingRequest:
    """Submit site cost estimates, trigger pricing engine, generate spreadsheet."""
    request = pricing_request_repository.get_or_raise(db, request_id)

    if request.status != PricingRequestStatus.ESTIMATING:
        raise ValueError(
            f"Can only submit estimates for Estimating requests, "
            f"current status is '{request.status.value}'"
        )

    # Validate the submitter is the assigned estimator or admin
    is_admin = _has_role(estimator_profile, UserRoleType.ADMIN)
    if not is_admin and request.estimator_id != estimator_profile.profile_id:
        raise NotAuthorizedError(
            "Only the assigned estimator or admin can submit estimates"
        )

    # Store site cost inputs as JSONB
    site_cost_data = {
        li.lot_number: li.model_dump(mode="json")
        for li in submission.lot_inputs
    }

    # Build pricing context from form_data
    form_data = request.form_data or {}
    estate = db.get(Estate, request.estate_id)

    ctx = PricingContext(
        estate_id=request.estate_id,
        stage_id=request.stage_id,
        brand=request.brand,
        suburb=estate.suburb if estate else "",
        postcode=estate.postcode if estate else "",
        bdm_profile_id=form_data.get("bdm_profile_id"),
        wholesale_group_name=form_data.get("wholesale_group"),
        is_kdrb=form_data.get("is_kdrb", False),
        is_10_90_deal=form_data.get("is_10_90_deal", False),
        holding_costs_apply=form_data.get("holding_costs_apply", False),
        developer_land_referrals=form_data.get("developer_land_referrals", False),
        building_crossover=form_data.get("building_crossover", False),
        shared_crossovers=form_data.get("shared_crossovers", False),
    )

    # Build lot pricing inputs from form_data lots + estimator site cost inputs
    lots_data = form_data.get("lots", [])
    lot_inputs: list[LotPricingInput] = []

    for lot_entry in lots_data:
        lot_number = str(lot_entry.get("lot_number", ""))
        site_input = site_cost_data.get(lot_number, {})

        # Look up lot record for dimensions
        lot_record = db.execute(
            select(StageLot).where(
                StageLot.stage_id == request.stage_id,
                StageLot.lot_number == lot_number,
            )
        ).scalars().first()

        lot_inputs.append(
            LotPricingInput(
                lot_number=lot_number,
                lot_frontage=Decimal(str(lot_record.frontage)) if lot_record else Decimal("0"),
                lot_depth=Decimal(str(lot_record.depth)) if lot_record else Decimal("0"),
                lot_size_sqm=Decimal(str(lot_record.size_sqm)) if lot_record else Decimal("0"),
                land_price=Decimal(str(lot_record.land_price)) if lot_record else Decimal("0"),
                corner_block=bool(lot_record.corner_block) if lot_record else False,
                orientation=str(lot_record.orientation) if lot_record else "N",
                house_name=str(lot_entry.get("house_type", "")),
                facade_name=str(lot_entry.get("facade_type", "")),
                garage_side=str(lot_entry.get("garage_side", "Left") or "Left"),
                fall_mm=int(site_input.get("fall_mm", 0)),
                fill_trees=bool(site_input.get("fill_trees", False)),
                easement_proximity_lhs=bool(site_input.get("easement_proximity_lhs", False)),
                easement_proximity_rhs=bool(site_input.get("easement_proximity_rhs", False)),
                retaining_lhs=bool(site_input.get("retaining_lhs", False)),
                retaining_rhs=bool(site_input.get("retaining_rhs", False)),
            )
        )

    # Run pricing engine
    breakdowns = calculate_batch(db, ctx, lot_inputs)
    price_breakdown_data = [bd.to_dict() for bd in breakdowns]

    # Generate spreadsheet
    generated_path = None
    try:
        template = get_template_for_brand(db, request.brand)
        if template is not None:
            generated_path = generate_pricing_spreadsheet(db, request, template)
    except Exception as exc:
        logger.warning("Spreadsheet generation failed: %s", exc)

    # Update request
    pricing_request_repository.update(
        db,
        request,
        site_cost_inputs=site_cost_data,
        price_breakdown=price_breakdown_data,
        estimated_at=datetime.now(timezone.utc),
        status=PricingRequestStatus.PRICED,
        generated_file_path=generated_path,
        pricing_engine_version="3.0",
    )

    return request


def get_request_detail(
    db: Session,
    request_id: int,
    user_profile: Profile,
) -> PricingRequestDetailRead:
    """Get enriched request detail with role filtering."""
    request = pricing_request_repository.get_or_raise(db, request_id)

    # Role filtering: sales/requester can only see own requests
    if not _is_admin_or_pricing(user_profile):
        if request.requester_id != user_profile.profile_id:
            raise NotAuthorizedError("You can only view your own pricing requests")

    # Enrich with names
    requester = db.get(Profile, request.requester_id)
    estate = db.get(Estate, request.estate_id)
    stage = db.get(EstateStage, request.stage_id)
    estimator = db.get(Profile, request.estimator_id) if request.estimator_id else None

    base = PricingRequestRead.model_validate(request).model_dump()
    return PricingRequestDetailRead(
        **base,
        requester_name=f"{requester.first_name} {requester.last_name}" if requester else None,
        estate_name=estate.estate_name if estate else None,
        stage_name=stage.name if stage else None,
        estimator_name=f"{estimator.first_name} {estimator.last_name}" if estimator else None,
    )


def list_requests(
    db: Session,
    user_profile: Profile,
    filters: dict,
    page: int = 1,
    size: int = 25,
) -> PaginatedResponse[PricingRequestRead]:
    """List pricing requests with role-based filtering."""
    requester_id = None
    if not _is_admin_or_pricing(user_profile):
        requester_id = user_profile.profile_id

    items, total = pricing_request_repository.list_paginated(
        db, filters, page, size, requester_id=requester_id
    )
    pages = math.ceil(total / size) if size else 0

    return PaginatedResponse[PricingRequestRead](
        items=[PricingRequestRead.model_validate(r) for r in items],
        total=total,
        page=page,
        size=size,
        pages=pages,
    )


def resubmit_request(
    db: Session,
    request_id: int,
    user_profile: Profile,
) -> dict[str, Any]:
    """Return form data from completed request for cloning as new draft."""
    request = pricing_request_repository.get_or_raise(db, request_id)

    if not _is_admin_or_pricing(user_profile):
        if request.requester_id != user_profile.profile_id:
            raise NotAuthorizedError("You can only resubmit your own pricing requests")

    return request.form_data


def delete_request(
    db: Session,
    request_id: int,
    user_profile: Profile,
) -> None:
    """Delete a pricing request. Owners can delete own; admin/pricing can delete any."""
    request = pricing_request_repository.get_or_raise(db, request_id)

    if not _is_admin_or_pricing(user_profile):
        if request.requester_id != user_profile.profile_id:
            raise NotAuthorizedError("You can only delete your own pricing requests")

    pricing_request_repository.delete(db, request_id)
