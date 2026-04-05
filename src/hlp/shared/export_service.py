"""CSV / XLSX exports for Land Search Interface results."""

from __future__ import annotations

import csv
import io
from collections.abc import Iterable
from datetime import datetime
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# Column order used by both CSV and XLSX exports.
EXPORT_HEADERS: list[str] = [
    "Estate",
    "Developer",
    "Region",
    "Suburb",
    "State",
    "Stage",
    "Lot Number",
    "Status",
    "Frontage (m)",
    "Depth (m)",
    "Size (m2)",
    "Corner",
    "Land Price",
    "Build Price",
    "Package Price",
    "Title Date",
    "Last Confirmed",
]

# Currency columns (1-indexed) for XLSX number formatting.
_CURRENCY_COLUMNS = (13, 14, 15)  # Land Price, Build Price, Package Price
_CURRENCY_FORMAT = '"$"#,##0.00'


def _status_value(status: Any) -> str:
    if status is None:
        return ""
    return getattr(status, "value", str(status))


def _decimal_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return f"{value:.2f}" if value % 1 else str(value)
    return str(value)


def _decimal_to_float(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    return value


def _strip_tz(value: Any) -> Any:
    """Excel does not support timezone-aware datetimes — drop tzinfo."""
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def _date_to_str(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _row_to_values(row: Any) -> list[Any]:
    """Extract a single result row's fields in EXPORT_HEADERS order.

    ``row`` is expected to expose the LotSearchResult fields (lot + joined
    estate/developer/region/stage data).
    """
    return [
        getattr(row, "estate_name", ""),
        getattr(row, "developer_name", ""),
        getattr(row, "region_name", "") or "",
        getattr(row, "estate_suburb", "") or "",
        getattr(row, "estate_state", "") or "",
        getattr(row, "stage_name", ""),
        row.lot_number,
        _status_value(row.status),
        row.frontage,
        row.depth,
        row.size_sqm,
        "Yes" if row.corner_block else "No",
        row.land_price,
        row.build_price,
        row.package_price,
        row.title_date,
        row.last_confirmed_date,
    ]


def export_lots_csv(rows: Iterable[Any]) -> bytes:
    """Render a list of LotSearchResult rows as CSV bytes (UTF-8)."""
    buffer = io.StringIO()
    writer = csv.writer(buffer, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
    writer.writerow(EXPORT_HEADERS)
    for row in rows:
        values = _row_to_values(row)
        writer.writerow(
            [
                values[0],
                values[1],
                values[2],
                values[3],
                values[4],
                values[5],
                values[6],
                values[7],
                _decimal_to_str(values[8]),
                _decimal_to_str(values[9]),
                _decimal_to_str(values[10]),
                values[11],
                _decimal_to_str(values[12]),
                _decimal_to_str(values[13]),
                _decimal_to_str(values[14]),
                _date_to_str(values[15]),
                _date_to_str(values[16]),
            ]
        )
    return buffer.getvalue().encode("utf-8")


def export_lots_xlsx(rows: Iterable[Any]) -> bytes:
    """Render a list of LotSearchResult rows as XLSX bytes.

    Formatting:
      - Bold header row, frozen at top
      - Auto-sized columns based on content width
      - Currency formatting on price columns
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Lots"

    # Header row.
    header_font = Font(bold=True)
    for col_idx, header in enumerate(EXPORT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A2"

    # Data rows.
    for r_offset, row in enumerate(rows, start=2):
        values = _row_to_values(row)
        # Convert Decimals to floats for numeric columns so Excel treats them as numbers.
        numeric_values = [
            values[0],
            values[1],
            values[2],
            values[3],
            values[4],
            values[5],
            values[6],
            values[7],
            _decimal_to_float(values[8]),
            _decimal_to_float(values[9]),
            _decimal_to_float(values[10]),
            values[11],
            _decimal_to_float(values[12]),
            _decimal_to_float(values[13]),
            _decimal_to_float(values[14]),
            _strip_tz(values[15]),
            _strip_tz(values[16]),
        ]
        for col_idx, val in enumerate(numeric_values, start=1):
            ws.cell(row=r_offset, column=col_idx, value=val)

    # Apply currency format to price columns.
    max_row = ws.max_row
    for col_idx in _CURRENCY_COLUMNS:
        col_letter = get_column_letter(col_idx)
        for row_idx in range(2, max_row + 1):
            ws[f"{col_letter}{row_idx}"].number_format = _CURRENCY_FORMAT

    # Auto-size columns (cap at 40 chars to avoid insane widths).
    for col_idx, header in enumerate(EXPORT_HEADERS, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header))
        for row_idx in range(2, max_row + 1):
            cell_val = ws[f"{col_letter}{row_idx}"].value
            if cell_val is None:
                continue
            cell_str = f"{cell_val}"
            if len(cell_str) > max_len:
                max_len = len(cell_str)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
