"""Spreadsheet generation service: inject pricing request data into brand template."""

from __future__ import annotations

import io
import logging
from pathlib import PurePosixPath
from typing import Any
from uuid import uuid4

from sqlalchemy.orm import Session

from hlp.models.pricing_request import PricingRequest
from hlp.models.pricing_template import PricingTemplate
from hlp.repositories import pricing_rule_repository
from hlp.shared.pricing_rule_service import evaluate_rules
from hlp.shared.storage_service import CATEGORY_GENERATED_SHEETS, get_storage_service

logger = logging.getLogger(__name__)


def _build_request_context(form_data: dict[str, Any]) -> dict[str, Any]:
    """Build a context dict for pricing rule condition evaluation."""
    context: dict[str, Any] = {}
    # Boolean toggles
    for key in (
        "is_kdrb",
        "is_10_90_deal",
        "developer_land_referrals",
        "building_crossover",
        "shared_crossovers",
    ):
        context[key] = form_data.get(key, False)

    # Wholesale group
    if form_data.get("wholesale_group"):
        context["wholesale_group"] = form_data["wholesale_group"]

    # Collect house types from lots
    lots = form_data.get("lots", [])
    context["house_types"] = list({lot.get("house_type", "") for lot in lots if lot.get("house_type")})

    # Corner block — true if any lot is a corner block
    context["corner_block"] = any(lot.get("corner_block", False) for lot in lots)

    # Custom house design
    context["custom_house_design"] = any(lot.get("custom_house_design", False) for lot in lots)

    return context


def generate_pricing_spreadsheet(
    db: Session,
    request: PricingRequest,
    template: PricingTemplate,
) -> str:
    """Generate a pricing spreadsheet by injecting request data into the brand template.

    Returns the stored file path.
    """
    from openpyxl import load_workbook

    storage = get_storage_service()

    # 1. Load template file from storage
    template_bytes = storage.read_file(template.file_path)

    # 2. Open with openpyxl (keep_vba for .xlsm)
    is_xlsm = template.file_path.endswith(".xlsm")
    wb = load_workbook(
        io.BytesIO(template_bytes),
        keep_vba=is_xlsm,
    )

    # 3. Get target worksheet
    if template.sheet_name in wb.sheetnames:
        ws = wb[template.sheet_name]
    else:
        ws = wb.active

    form_data = request.form_data or {}

    # 4. Inject header data using template.header_mappings
    header_mappings = template.header_mappings or {}
    header_values = _resolve_header_values(form_data, request)
    for field_name, cell_ref in header_mappings.items():
        if isinstance(cell_ref, dict) and "row" in cell_ref and "col" in cell_ref:
            value = header_values.get(field_name, "")
            ws.cell(row=cell_ref["row"], column=cell_ref["col"], value=value)

    # 5. Inject lot data starting at data_start_row
    lots = form_data.get("lots", [])
    column_mappings = template.column_mappings or {}
    for i, lot in enumerate(lots):
        row = template.data_start_row + i
        for field_name, col in column_mappings.items():
            if isinstance(col, int):
                value = lot.get(field_name, "")
                ws.cell(row=row, column=col, value=value)

    # 6. Evaluate pricing rules and inject into mapped cells
    context = _build_request_context(form_data)
    brand = request.brand

    global_rules = pricing_rule_repository.list_global_rules(db, brand)
    stage_rules = pricing_rule_repository.list_stage_rules(
        db, request.estate_id, request.stage_id
    )
    all_rules = list(global_rules) + list(stage_rules)
    applicable_rules = evaluate_rules(all_rules, context)

    for rule in applicable_rules:
        if rule.cell_row and rule.cell_col:
            ws.cell(row=rule.cell_row, column=rule.cell_col, value=rule.item_name)
        if rule.cost_cell_row and rule.cost_cell_col:
            ws.cell(
                row=rule.cost_cell_row,
                column=rule.cost_cell_col,
                value=float(rule.cost),
            )

    # 7. Save workbook to bytes
    output = io.BytesIO()
    wb.save(output)
    wb.close()
    content = output.getvalue()

    # 8. Store via StorageService
    ext = ".xlsm" if is_xlsm else ".xlsx"
    filename = f"pricing_request_{request.request_id}_{uuid4().hex}{ext}"
    stored_path, _ = storage.save_file(CATEGORY_GENERATED_SHEETS, filename, content)

    return stored_path


def _resolve_header_values(
    form_data: dict[str, Any], request: PricingRequest
) -> dict[str, Any]:
    """Build a mapping of header field names to their values for injection."""
    values: dict[str, Any] = {
        "estate_id": request.estate_id,
        "stage_id": request.stage_id,
        "brand": request.brand,
        "has_land_titled": form_data.get("has_land_titled", ""),
        "titling_when": form_data.get("titling_when", ""),
        "is_kdrb": form_data.get("is_kdrb", False),
        "is_10_90_deal": form_data.get("is_10_90_deal", False),
        "developer_land_referrals": form_data.get("developer_land_referrals", False),
        "building_crossover": form_data.get("building_crossover", False),
        "shared_crossovers": form_data.get("shared_crossovers", False),
        "side_easement": form_data.get("side_easement", ""),
        "rear_easement": form_data.get("rear_easement", ""),
        "bdm": form_data.get("bdm", ""),
        "wholesale_group": form_data.get("wholesale_group", ""),
        "notes": form_data.get("notes", ""),
    }
    return values
