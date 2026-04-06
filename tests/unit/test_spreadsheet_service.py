"""Unit tests: spreadsheet generation service."""

from __future__ import annotations

import io
from datetime import datetime, timezone
from decimal import Decimal
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook

from hlp.models.pricing_request import PricingRequest
from hlp.models.pricing_template import PricingTemplate
from hlp.shared.spreadsheet_service import (
    _build_request_context,
    generate_pricing_spreadsheet,
)


def _make_template_bytes(sheet_name: str = "Sheet1") -> bytes:
    """Create a minimal .xlsx in memory for testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    # Add some header cells
    ws.cell(row=1, column=1, value="Estate")
    ws.cell(row=1, column=2, value="Brand")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_template(
    template_id: int = 1,
    brand: str = "Hermitage Homes",
    file_path: str = "pricing-templates/test.xlsx",
    sheet_name: str = "Sheet1",
    data_start_row: int = 3,
) -> PricingTemplate:
    t = PricingTemplate()
    t.template_id = template_id
    t.brand = brand
    t.file_path = file_path
    t.sheet_name = sheet_name
    t.data_start_row = data_start_row
    t.header_mappings = {
        "brand": {"row": 1, "col": 2},
        "bdm": {"row": 2, "col": 2},
    }
    t.column_mappings = {
        "lot_number": 1,
        "house_type": 2,
        "facade_type": 3,
    }
    t.data_validations = {}
    return t


def _make_request(
    request_id: int = 1,
    brand: str = "Hermitage Homes",
    estate_id: int = 1,
    stage_id: int = 1,
) -> PricingRequest:
    r = PricingRequest()
    r.request_id = request_id
    r.requester_id = 1
    r.estate_id = estate_id
    r.stage_id = stage_id
    r.brand = brand
    r.status = "Pending"
    r.form_data = {
        "brand": brand,
        "bdm": "John Doe",
        "wholesale_group": "GroupA",
        "is_kdrb": True,
        "lots": [
            {"lot_number": "1", "house_type": "Alpine", "facade_type": "Traditional"},
            {"lot_number": "2", "house_type": "Baxter", "facade_type": "Modern"},
        ],
    }
    r.lot_numbers = ["1", "2"]
    return r


@patch("hlp.shared.spreadsheet_service.get_storage_service")
@patch("hlp.shared.spreadsheet_service.pricing_rule_repository")
def test_generate_spreadsheet_produces_valid_xlsx(mock_rules_repo, mock_storage):
    """Generated spreadsheet should be a valid xlsx that can be opened."""
    template_bytes = _make_template_bytes()
    storage_instance = MagicMock()
    storage_instance.read_file.return_value = template_bytes
    storage_instance.save_file.return_value = ("generated-sheets/test.xlsx", 1000)
    mock_storage.return_value = storage_instance
    mock_rules_repo.list_global_rules.return_value = []
    mock_rules_repo.list_stage_rules.return_value = []

    db = MagicMock()
    template = _make_template()
    request = _make_request()

    result = generate_pricing_spreadsheet(db, request, template)

    assert result == "generated-sheets/test.xlsx"
    storage_instance.save_file.assert_called_once()
    saved_content = storage_instance.save_file.call_args[0][2]
    # Verify the saved content is valid xlsx
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(saved_content))
    assert wb.active is not None
    wb.close()


@patch("hlp.shared.spreadsheet_service.get_storage_service")
@patch("hlp.shared.spreadsheet_service.pricing_rule_repository")
def test_header_injection_at_mapped_cells(mock_rules_repo, mock_storage):
    """Header values should be written to the mapped cell positions."""
    template_bytes = _make_template_bytes()
    storage_instance = MagicMock()
    storage_instance.read_file.return_value = template_bytes
    storage_instance.save_file.return_value = ("generated-sheets/test.xlsx", 1000)
    mock_storage.return_value = storage_instance
    mock_rules_repo.list_global_rules.return_value = []
    mock_rules_repo.list_stage_rules.return_value = []

    db = MagicMock()
    template = _make_template()
    request = _make_request()

    generate_pricing_spreadsheet(db, request, template)

    saved_content = storage_instance.save_file.call_args[0][2]
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(saved_content))
    ws = wb.active
    # Brand should be at row 1, col 2
    assert ws.cell(row=1, column=2).value == "Hermitage Homes"
    # BDM should be at row 2, col 2
    assert ws.cell(row=2, column=2).value == "John Doe"
    wb.close()


@patch("hlp.shared.spreadsheet_service.get_storage_service")
@patch("hlp.shared.spreadsheet_service.pricing_rule_repository")
def test_lot_data_starts_at_configured_row(mock_rules_repo, mock_storage):
    """Lot data should be written starting at the template's data_start_row."""
    template_bytes = _make_template_bytes()
    storage_instance = MagicMock()
    storage_instance.read_file.return_value = template_bytes
    storage_instance.save_file.return_value = ("generated-sheets/test.xlsx", 1000)
    mock_storage.return_value = storage_instance
    mock_rules_repo.list_global_rules.return_value = []
    mock_rules_repo.list_stage_rules.return_value = []

    db = MagicMock()
    template = _make_template(data_start_row=5)
    request = _make_request()

    generate_pricing_spreadsheet(db, request, template)

    saved_content = storage_instance.save_file.call_args[0][2]
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(saved_content))
    ws = wb.active
    # First lot at row 5
    assert ws.cell(row=5, column=1).value == "1"
    assert ws.cell(row=5, column=2).value == "Alpine"
    assert ws.cell(row=5, column=3).value == "Traditional"
    # Second lot at row 6
    assert ws.cell(row=6, column=1).value == "2"
    assert ws.cell(row=6, column=2).value == "Baxter"
    wb.close()


@patch("hlp.shared.spreadsheet_service.get_storage_service")
@patch("hlp.shared.spreadsheet_service.pricing_rule_repository")
def test_pricing_rules_injected_conditionally(mock_rules_repo, mock_storage):
    """Only applicable pricing rules should be injected into the spreadsheet."""
    template_bytes = _make_template_bytes()
    storage_instance = MagicMock()
    storage_instance.read_file.return_value = template_bytes
    storage_instance.save_file.return_value = ("generated-sheets/test.xlsx", 1000)
    mock_storage.return_value = storage_instance

    # Create mock rules: one with matching condition, one without
    rule_match = MagicMock()
    rule_match.condition = "is_kdrb"
    rule_match.item_name = "KDRB Bonus"
    rule_match.cost = Decimal("500.00")
    rule_match.cell_row = 10
    rule_match.cell_col = 1
    rule_match.cost_cell_row = 10
    rule_match.cost_cell_col = 2

    rule_no_match = MagicMock()
    rule_no_match.condition = "corner_block"
    rule_no_match.item_name = "Corner Premium"
    rule_no_match.cost = Decimal("1000.00")
    rule_no_match.cell_row = 11
    rule_no_match.cell_col = 1
    rule_no_match.cost_cell_row = 11
    rule_no_match.cost_cell_col = 2

    mock_rules_repo.list_global_rules.return_value = [rule_match, rule_no_match]
    mock_rules_repo.list_stage_rules.return_value = []

    db = MagicMock()
    template = _make_template()
    request = _make_request()
    # form_data has is_kdrb=True but no corner_block lots

    generate_pricing_spreadsheet(db, request, template)

    saved_content = storage_instance.save_file.call_args[0][2]
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(saved_content))
    ws = wb.active
    # KDRB rule should be injected
    assert ws.cell(row=10, column=1).value == "KDRB Bonus"
    assert ws.cell(row=10, column=2).value == 500.0
    # Corner block rule should NOT be injected
    assert ws.cell(row=11, column=1).value is None
    wb.close()


@patch("hlp.shared.spreadsheet_service.get_storage_service")
def test_missing_template_file_raises(mock_storage):
    """Should raise when template file is not found in storage."""
    storage_instance = MagicMock()
    storage_instance.read_file.side_effect = FileNotFoundError("Not found")
    mock_storage.return_value = storage_instance

    db = MagicMock()
    template = _make_template()
    request = _make_request()

    with pytest.raises(FileNotFoundError):
        generate_pricing_spreadsheet(db, request, template)


def test_build_request_context_extracts_toggles():
    """Context should include boolean toggles and aggregated fields."""
    form_data = {
        "is_kdrb": True,
        "building_crossover": True,
        "is_10_90_deal": False,
        "wholesale_group": "GroupA",
        "lots": [
            {"house_type": "Alpine", "corner_block": True, "custom_house_design": False},
            {"house_type": "Baxter", "corner_block": False, "custom_house_design": True},
        ],
    }
    ctx = _build_request_context(form_data)
    assert ctx["is_kdrb"] is True
    assert ctx["building_crossover"] is True
    assert ctx["is_10_90_deal"] is False
    assert ctx["wholesale_group"] == "GroupA"
    assert set(ctx["house_types"]) == {"Alpine", "Baxter"}
    assert ctx["corner_block"] is True
    assert ctx["custom_house_design"] is True
