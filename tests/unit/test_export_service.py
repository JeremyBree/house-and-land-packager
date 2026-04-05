"""Unit tests for CSV / XLSX export service (Sprint 3 LSI exports)."""
from __future__ import annotations

import csv
import io
from datetime import date, datetime
from decimal import Decimal
from types import SimpleNamespace

from openpyxl import load_workbook

from hlp.models.enums import LotStatus, Source
from hlp.shared.export_service import (
    EXPORT_HEADERS,
    export_lots_csv,
    export_lots_xlsx,
)


def _make_row(**overrides):
    """Factory: produce a row shape that mimics LotSearchResult / Row tuple."""
    defaults = dict(
        # Joined estate/developer/region/stage fields
        estate_name="Acacia Rise",
        developer_name="AlphaDev",
        region_name="Metro",
        estate_suburb="Tarneit",
        estate_state="VIC",
        stage_name="Stage 1",
        # Lot fields
        lot_id=1,
        stage_id=1,
        lot_number="A1-1",
        frontage=Decimal("12.50"),
        depth=Decimal("30.00"),
        size_sqm=Decimal("400.00"),
        corner_block=False,
        land_price=Decimal("350000.00"),
        build_price=Decimal("280000.00"),
        package_price=Decimal("630000.00"),
        status=LotStatus.AVAILABLE,
        title_date=date(2026, 6, 1),
        last_confirmed_date=datetime(2026, 3, 1, 12, 0, 0),
        source=Source.MANUAL,
        orientation=None,
        side_easement=None,
        rear_easement=None,
        street_name=None,
        substation=False,
    )
    defaults.update(overrides)
    return SimpleNamespace(**defaults)


# ------------------------------ CSV -----------------------------------------


def test_export_csv_returns_bytes():
    row = _make_row()
    content = export_lots_csv([row])
    assert isinstance(content, bytes)
    text = content.decode("utf-8")
    reader = list(csv.reader(io.StringIO(text)))
    assert reader[0] == EXPORT_HEADERS
    assert len(reader) == 2  # header + 1 data row


def test_export_csv_escapes_commas():
    row = _make_row(estate_name="Acacia, Rise", street_name="Main, St")
    content = export_lots_csv([row]).decode("utf-8")
    # The quoted field must be present (csv.writer quotes fields containing commas).
    assert '"Acacia, Rise"' in content
    # Parsing round-trip preserves the original value.
    reader = list(csv.reader(io.StringIO(content)))
    assert reader[1][0] == "Acacia, Rise"


def test_export_csv_handles_null_values():
    row = _make_row(
        land_price=None,
        build_price=None,
        package_price=None,
        frontage=None,
        depth=None,
        size_sqm=None,
        title_date=None,
        last_confirmed_date=None,
        region_name=None,
        estate_suburb=None,
        estate_state=None,
    )
    content = export_lots_csv([row]).decode("utf-8")
    reader = list(csv.reader(io.StringIO(content)))
    data = reader[1]
    # Region (idx 2), Suburb (3), State (4): empty strings
    assert data[2] == ""
    assert data[3] == ""
    assert data[4] == ""
    # Frontage/Depth/Size (8,9,10), Land/Build/Package (12,13,14), dates (15,16)
    for idx in (8, 9, 10, 12, 13, 14, 15, 16):
        assert data[idx] == ""


def test_export_csv_formats_currency_as_string():
    row = _make_row(land_price=Decimal("350000.00"))
    content = export_lots_csv([row]).decode("utf-8")
    reader = list(csv.reader(io.StringIO(content)))
    # Land Price column index = 12.
    assert reader[1][12] == "350000.00"


# ------------------------------ XLSX ----------------------------------------


def test_export_xlsx_returns_bytes():
    row = _make_row()
    content = export_lots_xlsx([row])
    assert isinstance(content, bytes)
    assert len(content) > 0
    # Valid xlsx can be opened.
    wb = load_workbook(io.BytesIO(content))
    assert "Lots" in wb.sheetnames


def test_export_xlsx_has_header_row():
    row = _make_row()
    content = export_lots_xlsx([row])
    wb = load_workbook(io.BytesIO(content))
    ws = wb["Lots"]
    for col_idx, header in enumerate(EXPORT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        assert cell.value == header
        assert cell.font.bold is True


def test_export_xlsx_column_count():
    row = _make_row()
    content = export_lots_xlsx([row])
    wb = load_workbook(io.BytesIO(content))
    ws = wb["Lots"]
    assert ws.max_column == 17
    assert len(EXPORT_HEADERS) == 17


def test_export_xlsx_empty_rows_still_has_header():
    content = export_lots_xlsx([])
    wb = load_workbook(io.BytesIO(content))
    ws = wb["Lots"]
    assert ws.cell(row=1, column=1).value == "Estate"
    assert ws.max_row == 1
