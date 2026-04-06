"""Unit tests for hlp.shared.template_service."""
from __future__ import annotations

import io
from unittest.mock import MagicMock, patch

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from hlp.shared import template_service


def _create_workbook_with_validations() -> bytes:
    """Create a minimal XLSX workbook with data validations and return as bytes."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"

    # Set headers in row 1
    ws.cell(row=1, column=1, value="house_type")
    ws.cell(row=1, column=2, value="facade_type")

    # Add data validations
    dv_house = DataValidation(type="list", formula1='"Access 18,Camden 25,Alpine 30"')
    dv_house.sqref = "A2:A100"
    ws.add_data_validation(dv_house)

    dv_facade = DataValidation(type="list", formula1='"Traditional,Modern,Classic"')
    dv_facade.sqref = "B2:B100"
    ws.add_data_validation(dv_facade)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _create_empty_workbook() -> bytes:
    """Create a minimal XLSX workbook with no data validations."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value="header")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- Test: extract data validations from xlsx ---


def test_extract_data_validations_from_xlsx(tmp_path):
    """Create a workbook with data validations, save to file, and extract."""
    content = _create_workbook_with_validations()
    file_path = tmp_path / "pricing-templates" / "test_template.xlsx"
    file_path.parent.mkdir(parents=True, exist_ok=True)
    file_path.write_bytes(content)

    # Patch get_settings so storage_base_path points to tmp_path
    with patch("hlp.shared.template_service.get_settings") as mock_settings:
        mock_settings.return_value.storage_base_path = str(tmp_path)
        result = template_service.extract_data_validations(
            "pricing-templates/test_template.xlsx"
        )

    assert "house_type" in result
    assert "facade_type" in result
    assert "Access 18" in result["house_type"]
    assert "Camden 25" in result["house_type"]
    assert "Alpine 30" in result["house_type"]
    assert "Traditional" in result["facade_type"]
    assert "Modern" in result["facade_type"]
    assert "Classic" in result["facade_type"]


# --- Test: empty workbook returns empty dict ---


def test_extract_data_validations_empty_workbook(tmp_path):
    """An empty workbook should return an empty validations dict."""
    content = _create_empty_workbook()
    file_path = tmp_path / "pricing-templates" / "empty.xlsx"
    file_path.parent.mkdir(parents=True, exist_ok=True)
    file_path.write_bytes(content)

    with patch("hlp.shared.template_service.get_settings") as mock_settings:
        mock_settings.return_value.storage_base_path = str(tmp_path)
        result = template_service.extract_data_validations(
            "pricing-templates/empty.xlsx"
        )

    assert result == {}


# --- Test: upload_template saves file ---


def test_upload_template_saves_file(tmp_path, db_session):
    """Verify that upload_template saves the file via _save_template_file."""
    content = _create_empty_workbook()

    with patch("hlp.shared.template_service.get_settings") as mock_settings:
        mock_settings.return_value.storage_base_path = str(tmp_path)
        template = template_service.upload_template(
            db_session, "Hermitage Homes", content, "test.xlsx"
        )

    assert template.brand == "Hermitage Homes"
    assert template.file_path.startswith("pricing-templates/")
    assert template.sheet_name == "Sheet1"
    assert template.data_start_row == 2

    # Verify file was actually saved
    import pathlib
    saved = tmp_path / pathlib.PurePosixPath(template.file_path)
    assert saved.exists()
    assert saved.read_bytes() == content


# --- Test: update_template_mappings partial ---


def test_update_template_mappings_partial(tmp_path, db_session):
    """Only the provided fields should be updated."""
    content = _create_empty_workbook()

    with patch("hlp.shared.template_service.get_settings") as mock_settings:
        mock_settings.return_value.storage_base_path = str(tmp_path)
        template = template_service.upload_template(
            db_session, "Kingsbridge Homes", content, "kb.xlsx"
        )
        db_session.flush()

        # Update only sheet_name
        updated = template_service.update_template_mappings(
            db_session,
            template.template_id,
            sheet_name="PricingSheet",
        )

    assert updated.sheet_name == "PricingSheet"
    assert updated.data_start_row == 2  # unchanged
