"""Integration tests: /api/lots/export/csv and /api/lots/export/xlsx."""
from __future__ import annotations

import io

from fastapi.testclient import TestClient
from openpyxl import load_workbook


# ------------------------ auth ----------------------------------------------


def test_export_unauthenticated_401(client: TestClient):
    r = client.post("/api/lots/export/csv", json={})
    assert r.status_code == 401
    r2 = client.post("/api/lots/export/xlsx", json={})
    assert r2.status_code == 401


# ------------------------ CSV -----------------------------------------------


def test_csv_export_returns_csv_content_type(
    client: TestClient, sales_headers, search_seed_data
):
    r = client.post("/api/lots/export/csv", json={}, headers=sales_headers)
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith("text/csv")
    cd = r.headers.get("content-disposition", "")
    assert "attachment" in cd
    assert ".csv" in cd


def test_csv_export_contains_header_row(
    client: TestClient, sales_headers, search_seed_data
):
    r = client.post("/api/lots/export/csv", json={}, headers=sales_headers)
    assert r.status_code == 200
    text = r.content.decode("utf-8")
    first_line = text.split("\r\n", 1)[0]
    assert first_line.startswith("Estate,Developer,Region")
    assert "Lot Number" in first_line
    assert "Land Price" in first_line


def test_csv_export_contains_filtered_data(
    client: TestClient, sales_headers, search_seed_data
):
    # Filter by status = Sold (6 lots in the seed data).
    r = client.post(
        "/api/lots/export/csv",
        json={"statuses": ["Sold"]},
        headers=sales_headers,
    )
    assert r.status_code == 200
    text = r.content.decode("utf-8")
    lines = [ln for ln in text.split("\r\n") if ln]
    # header + 6 data rows
    assert len(lines) == 7
    for line in lines[1:]:
        assert "Sold" in line


def test_csv_export_empty_result_returns_header_only(
    client: TestClient, sales_headers, search_seed_data
):
    # Impossible filter: price above any seeded value.
    r = client.post(
        "/api/lots/export/csv",
        json={"price_min": 10_000_000},
        headers=sales_headers,
    )
    assert r.status_code == 200
    text = r.content.decode("utf-8")
    lines = [ln for ln in text.split("\r\n") if ln]
    assert len(lines) == 1  # header only
    assert lines[0].startswith("Estate,Developer,Region")


# ------------------------ XLSX ----------------------------------------------


def test_xlsx_export_returns_xlsx_content_type(
    client: TestClient, sales_headers, search_seed_data
):
    r = client.post("/api/lots/export/xlsx", json={}, headers=sales_headers)
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    cd = r.headers.get("content-disposition", "")
    assert "attachment" in cd
    assert ".xlsx" in cd


def test_xlsx_export_is_valid_xlsx(
    client: TestClient, sales_headers, search_seed_data
):
    r = client.post("/api/lots/export/xlsx", json={}, headers=sales_headers)
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    # The export writes to sheet titled "Lots" (the active sheet).
    assert "Lots" in wb.sheetnames
    ws = wb["Lots"]
    # Header + 30 seeded lots.
    assert ws.max_row == 31
    assert ws.cell(row=1, column=1).value == "Estate"


# ------------------------ 413 oversized export ------------------------------


def test_export_over_5000_rows_returns_413(
    client: TestClient, sales_headers, monkeypatch
):
    """Mock the repository count to simulate an oversized export."""
    from hlp.api.routers import lot_search as lot_search_router

    def _fake_count(_db, _filters):
        return 5001

    monkeypatch.setattr(
        lot_search_router.lot_search_repository, "count", _fake_count
    )

    r = client.post("/api/lots/export/csv", json={}, headers=sales_headers)
    assert r.status_code == 413
    body = r.json()
    assert body["code"] == "export_too_large"
