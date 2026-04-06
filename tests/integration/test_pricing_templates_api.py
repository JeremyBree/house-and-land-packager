"""Integration tests: /api/pricing-templates endpoints."""
from __future__ import annotations

import io
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation


def _make_xlsx_bytes(with_validations: bool = False) -> bytes:
    """Create a minimal XLSX workbook and return as bytes."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.cell(row=1, column=1, value="house_type")
    ws.cell(row=1, column=2, value="facade_type")
    if with_validations:
        dv = DataValidation(type="list", formula1='"Access 18,Camden 25"')
        dv.sqref = "A2:A100"
        ws.add_data_validation(dv)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---- list templates -----------------------------------------------------------


def test_admin_lists_templates(
    client: TestClient, admin_headers, tmp_storage, seeded_users
):
    r = client.get("/api/pricing-templates", headers=admin_headers)
    assert r.status_code == 200
    assert isinstance(r.json(), list)


# ---- upload -------------------------------------------------------------------


def test_admin_uploads_template_multipart(
    client: TestClient, admin_headers, tmp_storage, seeded_users
):
    content = _make_xlsx_bytes()
    r = client.post(
        "/api/pricing-templates/hermitage/upload",
        files={"file": ("hermitage.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=admin_headers,
    )
    assert r.status_code == 201, r.text
    body = r.json()
    assert body["brand"] == "Hermitage Homes"
    assert body["sheet_name"] == "Sheet1"
    assert body["data_start_row"] == 2


# ---- get by brand slug -------------------------------------------------------


def test_get_template_by_brand_slug(
    client: TestClient, admin_headers, sales_headers, tmp_storage, seeded_users
):
    # Upload first
    content = _make_xlsx_bytes()
    client.post(
        "/api/pricing-templates/kingsbridge/upload",
        files={"file": ("kb.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=admin_headers,
    )
    # Any authenticated user can read by brand
    r = client.get("/api/pricing-templates/kingsbridge", headers=sales_headers)
    assert r.status_code == 200
    body = r.json()
    assert body["brand"] == "Kingsbridge Homes"


# ---- update mappings ----------------------------------------------------------


def test_update_mappings(
    client: TestClient, admin_headers, tmp_storage, seeded_users
):
    content = _make_xlsx_bytes()
    r_up = client.post(
        "/api/pricing-templates/hermitage/upload",
        files={"file": ("h.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=admin_headers,
    )
    template_id = r_up.json()["template_id"]

    r = client.patch(
        f"/api/pricing-templates/{template_id}/mappings",
        json={"sheet_name": "PricingSheet", "data_start_row": 5},
        headers=admin_headers,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["sheet_name"] == "PricingSheet"
    assert body["data_start_row"] == 5


# ---- get validations ----------------------------------------------------------


def test_get_validations_returns_dict(
    client: TestClient, admin_headers, sales_headers, tmp_storage, seeded_users
):
    content = _make_xlsx_bytes(with_validations=True)
    client.post(
        "/api/pricing-templates/hermitage/upload",
        files={"file": ("h.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=admin_headers,
    )
    r = client.get("/api/pricing-templates/hermitage/validations", headers=sales_headers)
    assert r.status_code == 200
    body = r.json()
    assert "validations" in body
    assert isinstance(body["validations"], dict)


# ---- non-admin upload 403 ----------------------------------------------------


def test_non_admin_upload_403(
    client: TestClient, sales_headers, tmp_storage, seeded_users
):
    content = _make_xlsx_bytes()
    r = client.post(
        "/api/pricing-templates/hermitage/upload",
        files={"file": ("h.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=sales_headers,
    )
    assert r.status_code == 403


# ---- invalid brand slug 404 --------------------------------------------------


def test_invalid_brand_slug_404(
    client: TestClient, admin_headers, seeded_users
):
    r = client.get("/api/pricing-templates/foobar", headers=admin_headers)
    assert r.status_code == 404


# ---- unauth 401 --------------------------------------------------------------


def test_unauth_401(client: TestClient, seeded_users):
    r = client.get("/api/pricing-templates")
    assert r.status_code == 401
