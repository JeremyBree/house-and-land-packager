"""Integration tests: /api/pricing-requests endpoints."""

from __future__ import annotations

import io
from unittest.mock import MagicMock, patch

from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy.orm import Session

from tests.integration.conftest import auth_header


def _make_template_bytes(sheet_name: str = "Sheet1") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _seed_template(session_factory, brand: str = "Hermitage Homes"):
    """Seed a pricing template for the given brand."""
    from hlp.models.pricing_template import PricingTemplate

    session: Session = session_factory()
    try:
        t = PricingTemplate(
            brand=brand,
            file_path="pricing-templates/test.xlsx",
            sheet_name="Sheet1",
            data_start_row=3,
            header_mappings={"brand": {"row": 1, "col": 1}},
            column_mappings={"lot_number": 1, "house_type": 2, "facade_type": 3},
            data_validations={},
        )
        session.add(t)
        session.commit()
        return t.template_id
    finally:
        session.close()


def _make_payload(estate_id: int, stage_id: int, brand: str = "Hermitage Homes"):
    payload = {
        "estate_id": estate_id,
        "stage_id": stage_id,
        "brand": brand,
        "has_land_titled": True,
        "is_kdrb": False,
        "lots": [
            {"lot_number": "A1", "house_type": "Alpine", "facade_type": "Traditional"},
        ],
    }
    if brand == "Hermitage Homes":
        payload["bdm"] = "John Doe"
        payload["wholesale_group"] = "GroupA"
    return payload


# ---- submit ------------------------------------------------------------------


def test_submit_creates_request_with_pending_status(
    client: TestClient,
    admin_headers,
    sample_stages_and_lots,
    _session_factory,
    tmp_storage,
):
    data = sample_stages_and_lots
    _seed_template(_session_factory)

    # Mock storage to return template bytes
    template_bytes = _make_template_bytes()
    with patch("hlp.shared.spreadsheet_service.get_storage_service") as mock_storage:
        storage_instance = MagicMock()
        storage_instance.read_file.return_value = template_bytes
        storage_instance.save_file.return_value = ("generated-sheets/test.xlsx", 1000)
        mock_storage.return_value = storage_instance

        payload = _make_payload(data["estate_id"], data["stage_a_id"])
        r = client.post("/api/pricing-requests", json=payload, headers=admin_headers)

    assert r.status_code == 201, r.text
    body = r.json()
    assert body["status"] == "Pending"
    assert body["brand"] == "Hermitage Homes"
    assert body["lot_numbers"] == ["A1"]
    # Spreadsheet generation now happens after estimator submits site costs,
    # so generated_file_path is None at submission time.
    assert body["generated_file_path"] is None
    assert "request_id" in body


def test_submit_with_clash_violations_returns_409(
    client: TestClient,
    admin_headers,
    sample_stages_and_lots,
    sample_clash_rules,
    _session_factory,
    tmp_storage,
):
    data = sample_stages_and_lots
    _seed_template(_session_factory)

    # A1 and A2 clash with each other (from sample_clash_rules)
    # Submit both with same design+facade → should trigger clash
    payload = _make_payload(data["estate_id"], data["stage_a_id"])
    payload["lots"] = [
        {"lot_number": "A1", "house_type": "Alpine", "facade_type": "Traditional"},
        {"lot_number": "A2", "house_type": "Alpine", "facade_type": "Traditional"},
    ]

    r = client.post("/api/pricing-requests", json=payload, headers=admin_headers)
    assert r.status_code == 409, r.text
    body = r.json()
    assert body["code"] == "clash_violation"
    assert "violations" in body
    assert len(body["violations"]) > 0


def test_submit_validates_hermitage_requires_bdm(
    client: TestClient,
    admin_headers,
    sample_stages_and_lots,
):
    data = sample_stages_and_lots
    payload = {
        "estate_id": data["estate_id"],
        "stage_id": data["stage_a_id"],
        "brand": "Hermitage Homes",
        "has_land_titled": True,
        "lots": [
            {"lot_number": "A1", "house_type": "Alpine", "facade_type": "Traditional"},
        ],
        # Missing bdm and wholesale_group
    }
    r = client.post("/api/pricing-requests", json=payload, headers=admin_headers)
    assert r.status_code == 422, r.text


# ---- list --------------------------------------------------------------------


def test_list_requests_role_filtered(
    client: TestClient,
    admin_headers,
    sales_headers,
    sample_stages_and_lots,
    _session_factory,
    seeded_users,
    tmp_storage,
):
    data = sample_stages_and_lots
    _seed_template(_session_factory)

    template_bytes = _make_template_bytes()
    with patch("hlp.shared.spreadsheet_service.get_storage_service") as mock_storage:
        storage_instance = MagicMock()
        storage_instance.read_file.return_value = template_bytes
        storage_instance.save_file.return_value = ("generated-sheets/test.xlsx", 1000)
        mock_storage.return_value = storage_instance

        # Admin creates a request
        payload = _make_payload(data["estate_id"], data["stage_a_id"])
        r = client.post("/api/pricing-requests", json=payload, headers=admin_headers)
        assert r.status_code == 201

    # Admin sees the request
    r = client.get("/api/pricing-requests", headers=admin_headers)
    assert r.status_code == 200
    assert r.json()["total"] >= 1

    # Sales user sees only own requests (none created by them)
    r = client.get("/api/pricing-requests", headers=sales_headers)
    assert r.status_code == 200
    assert r.json()["total"] == 0


def test_get_request_detail(
    client: TestClient,
    admin_headers,
    sample_stages_and_lots,
    _session_factory,
    tmp_storage,
):
    data = sample_stages_and_lots
    _seed_template(_session_factory)

    template_bytes = _make_template_bytes()
    with patch("hlp.shared.spreadsheet_service.get_storage_service") as mock_storage:
        storage_instance = MagicMock()
        storage_instance.read_file.return_value = template_bytes
        storage_instance.save_file.return_value = ("generated-sheets/test.xlsx", 1000)
        mock_storage.return_value = storage_instance

        payload = _make_payload(data["estate_id"], data["stage_a_id"])
        r = client.post("/api/pricing-requests", json=payload, headers=admin_headers)
        assert r.status_code == 201
        request_id = r.json()["request_id"]

    r = client.get(f"/api/pricing-requests/{request_id}", headers=admin_headers)
    assert r.status_code == 200
    body = r.json()
    assert body["request_id"] == request_id
    assert body["requester_name"] is not None
    assert body["estate_name"] is not None


def test_download_generated_sheet(
    client: TestClient,
    admin_headers,
    sample_stages_and_lots,
    _session_factory,
    tmp_storage,
):
    data = sample_stages_and_lots
    _seed_template(_session_factory)

    template_bytes = _make_template_bytes()
    with patch("hlp.shared.spreadsheet_service.get_storage_service") as mock_storage:
        storage_instance = MagicMock()
        storage_instance.read_file.return_value = template_bytes
        storage_instance.save_file.return_value = ("generated-sheets/test.xlsx", 1000)
        mock_storage.return_value = storage_instance

        payload = _make_payload(data["estate_id"], data["stage_a_id"])
        r = client.post("/api/pricing-requests", json=payload, headers=admin_headers)
        assert r.status_code == 201
        request_id = r.json()["request_id"]

    # Download should return 404 when generated_file_path is None (before estimation)
    r = client.get(
        f"/api/pricing-requests/{request_id}/download",
        headers=admin_headers,
    )
    assert r.status_code == 404


# ---- fulfil ------------------------------------------------------------------


def test_fulfil_uploads_completed_sheet_and_imports_packages(
    client: TestClient,
    admin_headers,
    sample_stages_and_lots,
    _session_factory,
    tmp_storage,
):
    data = sample_stages_and_lots
    _seed_template(_session_factory)

    template_bytes = _make_template_bytes()
    with patch("hlp.shared.spreadsheet_service.get_storage_service") as mock_storage:
        storage_instance = MagicMock()
        storage_instance.read_file.return_value = template_bytes
        storage_instance.save_file.return_value = ("generated-sheets/test.xlsx", 1000)
        mock_storage.return_value = storage_instance

        payload = _make_payload(data["estate_id"], data["stage_a_id"])
        r = client.post("/api/pricing-requests", json=payload, headers=admin_headers)
        assert r.status_code == 201
        request_id = r.json()["request_id"]

    # Create a completed sheet with lot data
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=3, column=1, value="A1")
    ws.cell(row=3, column=2, value="Alpine")
    ws.cell(row=3, column=3, value="Traditional")
    buf = io.BytesIO()
    wb.save(buf)
    completed_bytes = buf.getvalue()

    with patch("hlp.shared.pricing_request_service.get_storage_service") as mock_fulfil_storage:
        fulfil_instance = MagicMock()
        fulfil_instance.save_file.return_value = ("generated-sheets/completed.xlsx", 2000)
        mock_fulfil_storage.return_value = fulfil_instance

        r = client.post(
            f"/api/pricing-requests/{request_id}/fulfil",
            files={"file": ("completed.xlsx", completed_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=admin_headers,
        )
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["status"] == "Completed"
        assert body["completed_file_path"] is not None


def test_fulfil_creates_notification_for_requester(
    client: TestClient,
    admin_headers,
    sample_stages_and_lots,
    _session_factory,
    tmp_storage,
):
    data = sample_stages_and_lots
    _seed_template(_session_factory)

    template_bytes = _make_template_bytes()
    with patch("hlp.shared.spreadsheet_service.get_storage_service") as mock_storage:
        storage_instance = MagicMock()
        storage_instance.read_file.return_value = template_bytes
        storage_instance.save_file.return_value = ("generated-sheets/test.xlsx", 1000)
        mock_storage.return_value = storage_instance

        payload = _make_payload(data["estate_id"], data["stage_a_id"])
        r = client.post("/api/pricing-requests", json=payload, headers=admin_headers)
        assert r.status_code == 201
        request_id = r.json()["request_id"]

    completed_bytes = _make_template_bytes()
    with patch("hlp.shared.pricing_request_service.get_storage_service") as mock_fulfil_storage:
        fulfil_instance = MagicMock()
        fulfil_instance.save_file.return_value = ("generated-sheets/completed.xlsx", 2000)
        mock_fulfil_storage.return_value = fulfil_instance

        r = client.post(
            f"/api/pricing-requests/{request_id}/fulfil",
            files={"file": ("completed.xlsx", completed_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=admin_headers,
        )
        assert r.status_code == 200

    # Check notifications
    r = client.get("/api/notifications/unread-count", headers=admin_headers)
    assert r.status_code == 200
    assert r.json()["count"] >= 1


def test_fulfil_non_pricing_role_403(
    client: TestClient,
    sales_headers,
    sample_stages_and_lots,
    _session_factory,
    tmp_storage,
):
    data = sample_stages_and_lots
    _seed_template(_session_factory)

    template_bytes = _make_template_bytes()
    with patch("hlp.shared.spreadsheet_service.get_storage_service") as mock_storage:
        storage_instance = MagicMock()
        storage_instance.read_file.return_value = template_bytes
        storage_instance.save_file.return_value = ("generated-sheets/test.xlsx", 1000)
        mock_storage.return_value = storage_instance

        # Sales creates a request (this is allowed)
        payload = _make_payload(data["estate_id"], data["stage_a_id"])
        r = client.post("/api/pricing-requests", json=payload, headers=sales_headers)
        assert r.status_code == 201
        request_id = r.json()["request_id"]

    # Sales tries to fulfil — should be 403
    completed_bytes = _make_template_bytes()
    r = client.post(
        f"/api/pricing-requests/{request_id}/fulfil",
        files={"file": ("completed.xlsx", completed_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=sales_headers,
    )
    assert r.status_code == 403


# ---- resubmit ----------------------------------------------------------------


def test_resubmit_returns_form_data(
    client: TestClient,
    admin_headers,
    sample_stages_and_lots,
    _session_factory,
    tmp_storage,
):
    data = sample_stages_and_lots
    _seed_template(_session_factory)

    template_bytes = _make_template_bytes()
    with patch("hlp.shared.spreadsheet_service.get_storage_service") as mock_storage:
        storage_instance = MagicMock()
        storage_instance.read_file.return_value = template_bytes
        storage_instance.save_file.return_value = ("generated-sheets/test.xlsx", 1000)
        mock_storage.return_value = storage_instance

        payload = _make_payload(data["estate_id"], data["stage_a_id"])
        r = client.post("/api/pricing-requests", json=payload, headers=admin_headers)
        assert r.status_code == 201
        request_id = r.json()["request_id"]

    r = client.post(
        f"/api/pricing-requests/{request_id}/resubmit", headers=admin_headers
    )
    assert r.status_code == 200
    body = r.json()
    assert body["brand"] == "Hermitage Homes"
    assert body["lots"] is not None


# ---- delete ------------------------------------------------------------------


def test_delete_own_request_204(
    client: TestClient,
    admin_headers,
    sample_stages_and_lots,
    _session_factory,
    tmp_storage,
):
    data = sample_stages_and_lots
    _seed_template(_session_factory)

    template_bytes = _make_template_bytes()
    with patch("hlp.shared.spreadsheet_service.get_storage_service") as mock_storage:
        storage_instance = MagicMock()
        storage_instance.read_file.return_value = template_bytes
        storage_instance.save_file.return_value = ("generated-sheets/test.xlsx", 1000)
        mock_storage.return_value = storage_instance

        payload = _make_payload(data["estate_id"], data["stage_a_id"])
        r = client.post("/api/pricing-requests", json=payload, headers=admin_headers)
        assert r.status_code == 201
        request_id = r.json()["request_id"]

    r = client.delete(
        f"/api/pricing-requests/{request_id}", headers=admin_headers
    )
    assert r.status_code == 204


def test_delete_others_request_403_for_sales(
    client: TestClient,
    admin_headers,
    sales_headers,
    sample_stages_and_lots,
    _session_factory,
    tmp_storage,
):
    data = sample_stages_and_lots
    _seed_template(_session_factory)

    template_bytes = _make_template_bytes()
    with patch("hlp.shared.spreadsheet_service.get_storage_service") as mock_storage:
        storage_instance = MagicMock()
        storage_instance.read_file.return_value = template_bytes
        storage_instance.save_file.return_value = ("generated-sheets/test.xlsx", 1000)
        mock_storage.return_value = storage_instance

        # Admin creates request
        payload = _make_payload(data["estate_id"], data["stage_a_id"])
        r = client.post("/api/pricing-requests", json=payload, headers=admin_headers)
        assert r.status_code == 201
        request_id = r.json()["request_id"]

    # Sales tries to delete — should be 403
    r = client.delete(
        f"/api/pricing-requests/{request_id}", headers=sales_headers
    )
    assert r.status_code == 403


def test_admin_can_delete_any_request(
    client: TestClient,
    admin_headers,
    sales_headers,
    sample_stages_and_lots,
    _session_factory,
    tmp_storage,
):
    data = sample_stages_and_lots
    _seed_template(_session_factory)

    template_bytes = _make_template_bytes()
    with patch("hlp.shared.spreadsheet_service.get_storage_service") as mock_storage:
        storage_instance = MagicMock()
        storage_instance.read_file.return_value = template_bytes
        storage_instance.save_file.return_value = ("generated-sheets/test.xlsx", 1000)
        mock_storage.return_value = storage_instance

        # Sales creates request
        payload = _make_payload(data["estate_id"], data["stage_a_id"])
        r = client.post("/api/pricing-requests", json=payload, headers=sales_headers)
        assert r.status_code == 201
        request_id = r.json()["request_id"]

    # Admin can delete
    r = client.delete(
        f"/api/pricing-requests/{request_id}", headers=admin_headers
    )
    assert r.status_code == 204


def test_list_filter_by_status_and_brand(
    client: TestClient,
    admin_headers,
    sample_stages_and_lots,
    _session_factory,
    tmp_storage,
):
    data = sample_stages_and_lots
    _seed_template(_session_factory)

    template_bytes = _make_template_bytes()
    with patch("hlp.shared.spreadsheet_service.get_storage_service") as mock_storage:
        storage_instance = MagicMock()
        storage_instance.read_file.return_value = template_bytes
        storage_instance.save_file.return_value = ("generated-sheets/test.xlsx", 1000)
        mock_storage.return_value = storage_instance

        payload = _make_payload(data["estate_id"], data["stage_a_id"])
        r = client.post("/api/pricing-requests", json=payload, headers=admin_headers)
        assert r.status_code == 201

    # Filter by status
    r = client.get(
        "/api/pricing-requests?status=Pending", headers=admin_headers
    )
    assert r.status_code == 200
    assert r.json()["total"] >= 1

    # Filter by brand
    r = client.get(
        "/api/pricing-requests?brand=Hermitage+Homes", headers=admin_headers
    )
    assert r.status_code == 200
    assert r.json()["total"] >= 1

    # Filter by non-existent brand
    r = client.get(
        "/api/pricing-requests?brand=Unknown", headers=admin_headers
    )
    assert r.status_code == 200
    assert r.json()["total"] == 0


def test_unauth_401(client: TestClient):
    r = client.get("/api/pricing-requests")
    assert r.status_code == 401

    r = client.post("/api/pricing-requests", json={})
    assert r.status_code == 401
